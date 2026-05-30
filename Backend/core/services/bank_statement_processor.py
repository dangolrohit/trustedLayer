import logging
import re
from collections import defaultdict
from datetime import date as dt_date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from statistics import mean, pstdev
from uuid import uuid4

import pdfplumber
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb_workbook
import xlrd
from django.conf import settings
from django.core.exceptions import ValidationError
from django.utils import timezone

from core.models import BankStatement, BehavioralData
from core.utils.supabase_client import create_signed_url, download_private_file, upload_private_file

logger = logging.getLogger(__name__)

MONEY_PATTERN = r"[-+]?(?:Rs\.?|NPR)?\s?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?|[-+]?\d+(?:\.\d{1,2})?"
DATE_PATTERN = r"(?P<date>\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})"
OPENXML_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xlam"}
XLS_EXTENSIONS = {".xls", ".xlt", ".xla"}
XLSB_EXTENSIONS = {".xlsb"}
EXCEL_EXTENSIONS = OPENXML_EXTENSIONS | XLS_EXTENSIONS | XLSB_EXTENSIONS


def detect_statement_type(uploaded_file) -> str:
    name = (uploaded_file.name or "").lower()
    if name.endswith(".pdf") or uploaded_file.content_type == "application/pdf":
        return "pdf"
    if any(name.endswith(ext) for ext in EXCEL_EXTENSIONS):
        return "excel"
    raise ValidationError(
        "Only PDF or Excel bank statements are supported "
        "(.pdf, .xlsx, .xls, .xlsm, .xlsb, .xltx, .xltm)."
    )


def validate_statement_upload(uploaded_file):
    if uploaded_file.size > settings.MAX_BANK_STATEMENT_BYTES:
        raise ValidationError("Bank statement must be 5MB or smaller.")
    return detect_statement_type(uploaded_file)


def extract_pdf_text(uploaded_file) -> str:
    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()
    text_parts = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text:
                text_parts.append(text)
    return "\n".join(text_parts).strip()


def _to_decimal(raw) -> Decimal:
    if raw is None:
        return Decimal("0")
    if isinstance(raw, Decimal):
        return raw
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    cleaned = re.sub(r"[^\d\-.]", "", str(raw))
    if not cleaned:
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")


def _normalize_date(raw):
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, dt_date):
        return raw.isoformat()
    if isinstance(raw, (int, float)):
        # Excel serial date conversion base (Windows epoch).
        try:
            excel_base = datetime(1899, 12, 30)
            return (excel_base + timedelta(days=float(raw))).date().isoformat()
        except (OverflowError, ValueError):
            return None
    cleaned = (raw or "").strip().replace("/", "-")
    formats = (
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%d-%m-%y",
        "%m-%d-%y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(cleaned, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _extract_excel_rows(uploaded_file, extension: str):
    if extension in OPENXML_EXTENSIONS:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        workbook.close()
        return rows

    if extension in XLS_EXTENSIONS:
        uploaded_file.seek(0)
        workbook = xlrd.open_workbook(file_contents=uploaded_file.read())
        sheet = workbook.sheet_by_index(0)
        rows = []
        for row_index in range(sheet.nrows):
            row = []
            for col_index in range(sheet.ncols):
                cell = sheet.cell(row_index, col_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, workbook.datemode)
                row.append(value)
            rows.append(tuple(row))
        return rows

    if extension in XLSB_EXTENSIONS:
        uploaded_file.seek(0)
        workbook_bytes = uploaded_file.read()
        rows = []
        with open_xlsb_workbook(BytesIO(workbook_bytes)) as workbook:
            first_sheet_name = workbook.sheets[0]
            with workbook.get_sheet(first_sheet_name) as sheet:
                for row in sheet.rows():
                    rows.append(tuple(cell.v for cell in row))
        return rows

    return []


def parse_excel_transactions(uploaded_file) -> list[dict]:
    extension = "." + (uploaded_file.name or "").lower().rsplit(".", 1)[-1] if "." in (uploaded_file.name or "") else ""
    rows = _extract_excel_rows(uploaded_file, extension)
    if not rows:
        return []

    headers = {}
    header_row_index = 0
    aliases = {
        "date": {"date", "txn date", "transaction date", "value date", "posting date"},
        "description": {"description", "particulars", "narration", "details", "remarks", "remark"},
        "debit": {"debit", "withdrawal", "dr", "debits"},
        "credit": {"credit", "deposit", "cr", "credits"},
        "amount": {"amount", "txn amount", "transaction amount"},
        "balance": {"balance", "closing balance", "available balance"},
    }

    for index, row in enumerate(rows[:20]):
        normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if not normalized:
            continue
        row_map = {}
        for idx, value in enumerate(normalized):
            for key, words in aliases.items():
                if value in words:
                    row_map[key] = idx
        if "date" in row_map and ("amount" in row_map or "debit" in row_map or "credit" in row_map):
            headers = row_map
            header_row_index = index
            break

    if not headers:
        return []

    transactions = []
    for row in rows[header_row_index + 1 :]:
        if not row:
            continue
        raw_date = row[headers["date"]] if headers.get("date") is not None and len(row) > headers["date"] else None
        tx_date = _normalize_date(raw_date)
        if not tx_date:
            continue

        description = ""
        if headers.get("description") is not None and len(row) > headers["description"]:
            description = str(row[headers["description"]] or "").strip()

        debit = Decimal("0")
        credit = Decimal("0")
        balance = Decimal("0")

        if headers.get("debit") is not None and len(row) > headers["debit"]:
            debit = abs(_to_decimal(row[headers["debit"]]))
        if headers.get("credit") is not None and len(row) > headers["credit"]:
            credit = abs(_to_decimal(row[headers["credit"]]))
        if headers.get("balance") is not None and len(row) > headers["balance"]:
            balance = _to_decimal(row[headers["balance"]])

        if debit == 0 and credit == 0 and headers.get("amount") is not None and len(row) > headers["amount"]:
            amount = _to_decimal(row[headers["amount"]])
            if amount < 0:
                debit = abs(amount)
            else:
                credit = abs(amount)

        if not description:
            description = "Statement transaction"

        transactions.append(
            {
                "date": tx_date,
                "description": description[:255],
                "credit": float(credit),
                "debit": float(debit),
                "balance": float(balance),
            }
        )

    return transactions


def serialize_transactions_as_text(transactions: list[dict]) -> str:
    if not transactions:
        return ""
    lines = []
    for tx in transactions:
        lines.append(
            f"{tx['date']} | {tx.get('description', '')} | CR {tx.get('credit', 0)} | "
            f"DR {tx.get('debit', 0)} | BAL {tx.get('balance', 0)}"
        )
    return "\n".join(lines)


def parse_transactions(extracted_text: str) -> list[dict]:
    transactions = []
    for line in extracted_text.splitlines():
        compact = " ".join(line.split())
        if not compact:
            continue
        date_match = re.search(DATE_PATTERN, compact)
        if not date_match:
            continue

        amounts = re.findall(MONEY_PATTERN, compact)
        numeric_amounts = [_to_decimal(item) for item in amounts if re.search(r"\d", item)]
        if len(numeric_amounts) < 2:
            continue

        tx_date = _normalize_date(date_match.group("date"))
        if not tx_date:
            continue

        balance = numeric_amounts[-1]
        debit = Decimal("0")
        credit = Decimal("0")
        if len(numeric_amounts) >= 3:
            debit = abs(numeric_amounts[-3])
            credit = abs(numeric_amounts[-2])
        else:
            amount = numeric_amounts[-2]
            if any(token in compact.lower() for token in ["cr", "credit", "deposit", "received"]):
                credit = abs(amount)
            else:
                debit = abs(amount)

        description = compact[date_match.end() :]
        description = re.sub(MONEY_PATTERN, "", description).strip(" -|")
        transactions.append(
            {
                "date": tx_date,
                "description": description[:255],
                "credit": float(credit),
                "debit": float(debit),
                "balance": float(balance),
            }
        )
    return transactions


def analyze_transactions(transactions: list[dict]) -> dict:
    if not transactions:
        return {
            "monthly_income": 0,
            "avg_balance": 0,
            "consistency_score": 0,
            "volatility": 100,
            "bounced_count": 0,
            "income_expense_ratio": 0,
            "transaction_count": 0,
            "bank_behavior_score": 0,
        }

    monthly_income = defaultdict(float)
    balances = []
    total_credit = 0.0
    total_debit = 0.0
    bounced_count = 0

    for tx in transactions:
        month = tx["date"][:7]
        credit = float(tx.get("credit") or 0)
        debit = float(tx.get("debit") or 0)
        balance = float(tx.get("balance") or 0)
        total_credit += credit
        total_debit += debit
        monthly_income[month] += credit
        balances.append(balance)
        text = str(tx.get("description", "")).lower()
        if any(word in text for word in ["bounce", "returned", "insufficient", "dishonour"]):
            bounced_count += 1

    income_values = list(monthly_income.values()) or [0.0]
    avg_monthly_income = mean(income_values)
    income_volatility = (pstdev(income_values) / avg_monthly_income) if avg_monthly_income else 1
    consistency_score = max(0, min(100, round(100 - (income_volatility * 100))))
    avg_balance = mean(balances) if balances else 0
    balance_volatility = (pstdev(balances) / avg_balance) if avg_balance > 0 and len(balances) > 1 else 0
    income_expense_ratio = total_credit / total_debit if total_debit else total_credit

    base = 45
    base += min(25, avg_monthly_income / 4000)
    base += consistency_score * 0.25
    base += min(10, max(0, income_expense_ratio - 1) * 10)
    base -= min(25, balance_volatility * 25)
    base -= bounced_count * 8
    bank_behavior_score = round(max(0, min(100, base)))

    return {
        "monthly_income": round(avg_monthly_income, 2),
        "avg_balance": round(avg_balance, 2),
        "consistency_score": consistency_score,
        "volatility": round(balance_volatility * 100, 2),
        "bounced_count": bounced_count,
        "income_expense_ratio": round(income_expense_ratio, 2),
        "transaction_count": len(transactions),
        "months_observed": len(monthly_income),
        "total_credit": round(total_credit, 2),
        "total_debit": round(total_debit, 2),
        "bank_behavior_score": bank_behavior_score,
    }


class BankStatementProcessor:
    @staticmethod
    def process_upload(merchant, uploaded_file) -> BankStatement:
        statement_type = validate_statement_upload(uploaded_file)
        storage_path = (
            f"merchants/{merchant.id}/{timezone.now().date().isoformat()}/"
            f"{uuid4()}-{uploaded_file.name}"
        )

        if statement_type == "pdf":
            content_type = "application/pdf"
        else:
            content_type = uploaded_file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        uploaded_file.seek(0)
        upload_private_file(uploaded_file, storage_path, content_type)

        stored_bytes = download_private_file(storage_path)
        stored_file = SimpleUploadedFile(uploaded_file.name, stored_bytes, content_type=content_type)

        if statement_type == "pdf":
            extracted_text = extract_pdf_text(stored_file)
            parsed_transactions = parse_transactions(extracted_text)
        else:
            parsed_transactions = parse_excel_transactions(stored_file)
            extracted_text = serialize_transactions_as_text(parsed_transactions)

        analysis_summary = analyze_transactions(parsed_transactions)
        signed_url = create_signed_url(storage_path)

        statement = BankStatement.objects.create(
            merchant=merchant,
            file_path=storage_path,
            file_url=signed_url,
            extracted_text=extracted_text,
            parsed_transactions=parsed_transactions,
            analysis_summary=analysis_summary,
        )

        if parsed_transactions:
            dates = [datetime.fromisoformat(tx["date"]).date() for tx in parsed_transactions]
            BehavioralData.objects.create(
                merchant=merchant,
                data_type=BehavioralData.DataTypes.WALLET,
                metrics_json={
                    "source": "bank_statement",
                    "bank_statement_id": statement.id,
                    **analysis_summary,
                },
                period_start=min(dates),
                period_end=max(dates),
            )
        return statement
